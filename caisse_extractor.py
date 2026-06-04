# ==============================================================
#  extractors/caisse_extractor.py
#  Lit les fichiers Caisse W et Caisse B.
#  La date de référence pour le cross-check est extraite
#  du champ LIBELLÉ (ex: "CA DU 01/01/2026") en priorité,
#  avec fallback sur la colonne DATE si le libellé n'a pas de date.
# ==============================================================

import sys, os, re, datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from extractors.utils import nettoyer_valeur, lire_date, extraire_mois_annee
from config.mapping import CAISSE_COL_CODE, CAISSE_COL_DEBIT, CAISSE_COL_CREDIT, CAISSE_CODES_CA


def _extraire_date_libelle(libelle: str, annee_feuille: int) -> datetime.date | None:
    """
    Extrait la date depuis un libellé de caisse.

    Formats gérés :
      "CA DU 01/01/2026"          → 01/01/2026
      "CA DU 01/10"               → 01/10/<annee_feuille>
      "CHIFFRE D'AFFAIRE 20/06/2023 ESPECE" → 20/06/2023
      "CHIFFRE D'AFFAIRE DU 15/06 ESPECE"   → 15/06/<annee_feuille>

    Retourne None si aucune date trouvée ou si la date est invalide.
    """
    if not libelle or not isinstance(libelle, str):
        return None
    lib = libelle.upper().strip()

    # 1. Date complète dd/mm/yyyy
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})', lib)
    if m:
        j, mo, an = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # Si l année du libelle est erronee (ex: 2021 dans une feuille 2024),
        # on utilise l annee de la feuille pour corriger automatiquement
        if abs(an - annee_feuille) > 1:
            an = annee_feuille
        try:
            return datetime.date(an, mo, j)
        except ValueError:
            pass

    # 2. Date partielle dd/mm (sans année) → utiliser l'année de la feuille
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})(?![/\-\.]\d)', lib)
    if m:
        j, mo = int(m.group(1)), int(m.group(2))
        if 1 <= j <= 31 and 1 <= mo <= 12:
            try:
                return datetime.date(annee_feuille, mo, j)
            except ValueError:
                pass

    return None


def _trouver_entete_caisse(lignes: list) -> tuple:
    """Détecte la ligne d'en-tête et retourne (index, dict_colonnes)."""
    mots_code   = [c.upper() for c in CAISSE_COL_CODE]
    mots_debit  = [c.upper() for c in CAISSE_COL_DEBIT]
    mots_credit = [c.upper() for c in CAISSE_COL_CREDIT]

    for i, ligne in enumerate(lignes[:20]):
        vals = [str(c).upper().strip() for c in ligne if c is not None]
        if 'DATE' not in vals:
            continue
        if not any(m in vals for m in mots_debit):
            continue

        colonnes = {}
        for j, cell in enumerate(ligne):
            if cell is None:
                continue
            cle = str(cell).upper().strip()
            if cle in mots_code and 'code' not in colonnes:
                colonnes['code'] = j
            elif cle == 'DATE':
                colonnes['date'] = j
            elif cle == 'FNR' and 'fnr' not in colonnes:
                colonnes['fnr'] = j
            elif cle in ('LIBELLÉS','LIBELLE','LIBELLÉ','LIBELLES') and 'libelle' not in colonnes:
                colonnes['libelle'] = j
            elif cle in mots_debit and 'debit' not in colonnes:
                colonnes['debit'] = j
            elif cle in mots_credit and 'credit' not in colonnes:
                colonnes['credit'] = j
            elif cle == 'SOLDE' and 'solde' not in colonnes:
                colonnes['solde'] = j

        if 'date' in colonnes and 'debit' in colonnes:
            return i, colonnes

    return None, {}


def _est_entree_ca(code: str, fnr: str, libelle: str) -> bool:
    """
    Retourne True UNIQUEMENT si cette ligne est une entree CA especes journaliere.
    On accepte seulement les libelles du type "CA DU dd/mm/yyyy" ou "CA DU dd/mm".
    On exclut les tournois, ventes de materiel, regularisations, etc.
    """
    import re
    lib = libelle.upper().strip()
    
    # Pattern strict : "CA DU dd/mm/yyyy" ou "CA DU dd/mm" 
    if re.match(r'CA DU \d{1,2}/\d{1,2}', lib):
        return True
    
    # Pattern caisse W : "CHIFFRE D'AFFAIRE DU dd/mm"
    if re.match(r'CHIFFRE D[\'\s]?AFFAIRE[S]?\s+DU \d{1,2}/\d{1,2}', lib):
        return True
    
    # FNR = CA (caisse W)
    if fnr.upper().strip() == 'CA':
        if re.search(r'\d{1,2}/\d{1,2}', lib):
            return True
    
    return False


def _lire_feuille_caisse(ws, nom_feuille: str, id_caisse: str) -> list:
    """Parse une feuille de caisse et retourne les mouvements CA."""
    lignes = list(ws.iter_rows(values_only=True))
    idx_entete, colonnes = _trouver_entete_caisse(lignes)

    if idx_entete is None:
        return []

    # Extraire l'année depuis le nom de la feuille pour compléter les dates partielles
    my = extraire_mois_annee(nom_feuille)
    annee_feuille = my[1] if my else datetime.date.today().year

    mouvements = []
    for ligne in lignes[idx_entete + 1:]:
        if not ligne or len(ligne) <= max(colonnes.values(), default=0):
            continue

        def get(nom):
            if nom in colonnes and colonnes[nom] < len(ligne):
                return ligne[colonnes[nom]]
            return None

        code    = str(get('code')    or '').strip()
        fnr     = str(get('fnr')     or '').strip()
        libelle = str(get('libelle') or '').strip()
        debit   = nettoyer_valeur(get('debit'))
        credit  = nettoyer_valeur(get('credit'))
        solde   = nettoyer_valeur(get('solde'))

        # Vérifier si c'est une entrée CA
        if not _est_entree_ca(code, fnr, libelle):
            continue
        if debit is None or debit <= 0:
            continue

        # ── Clé de date : libellé en priorité, puis colonne DATE ──
        date_depuis_libelle = _extraire_date_libelle(libelle, annee_feuille)
        date_depuis_col     = lire_date(get('date'))

        # Choisir la meilleure date :
        # - Si le libellé donne une date valide → on l'utilise (c'est la date réelle du CA)
        # - Sinon on prend la colonne DATE
        if date_depuis_libelle is not None:
            date_val = date_depuis_libelle
            source_date = 'libelle'
        elif date_depuis_col is not None:
            date_val = date_depuis_col
            source_date = 'colonne'
        else:
            continue  # Aucune date trouvable → ignorer

        mouvements.append({
            'caisse':        id_caisse,
            'feuille':       nom_feuille,
            'date':          date_val,
            'date_colonne':  date_depuis_col,
            'date_libelle':  date_depuis_libelle,
            'source_date':   source_date,
            'code':          code,
            'fnr':           fnr,
            'libelle':       libelle,
            'debit':         debit,
            'credit':        credit,
            'solde':         solde,
            'est_entree_ca': True,
        })

    return mouvements


def extraire_caisse(chemin_fichier: str, id_caisse: str) -> dict:
    """
    Extrait toutes les entrées CA d'un fichier caisse.
    La date utilisée pour le cross-check provient du LIBELLÉ
    (ex: "CA DU 01/01/2026") en priorité.
    """
    print(f'\n  Lecture Caisse {id_caisse}...')
    print(f'  Fichier : {os.path.basename(chemin_fichier)}')

    try:
        wb = openpyxl.load_workbook(chemin_fichier, read_only=True, data_only=True)
    except FileNotFoundError:
        return {'mouvements': [], 'entrees_ca': [], 'nb_mouvements': 0,
                'nb_entrees_ca': 0, 'erreurs': [f'Fichier introuvable : {chemin_fichier}']}
    except Exception as e:
        return {'mouvements': [], 'entrees_ca': [], 'nb_mouvements': 0,
                'nb_entrees_ca': 0, 'erreurs': [f'Impossible d ouvrir : {e}']}

    tous = []
    erreurs = []

    for nom_feuille in wb.sheetnames:
        try:
            ws = wb[nom_feuille]
            mvts = _lire_feuille_caisse(ws, nom_feuille, id_caisse)
            if mvts:
                nb_lib = sum(1 for m in mvts if m['source_date'] == 'libelle')
                nb_col = sum(1 for m in mvts if m['source_date'] == 'colonne')
                print(f'  ✅ {nom_feuille} → {len(mvts)} entrées CA '
                      f'(date libellé={nb_lib}, date colonne={nb_col})')
                tous.extend(mvts)
        except Exception as e:
            msg = f'Erreur {nom_feuille}: {e}'
            erreurs.append(msg)

    wb.close()
    print(f'  → Total Caisse {id_caisse} : {len(tous)} entrées CA')

    return {
        'mouvements':    tous,
        'entrees_ca':    tous,
        'nb_mouvements': len(tous),
        'nb_entrees_ca': len(tous),
        'erreurs':       erreurs,
    }
