# ==============================================================
#  extractors/ca_extractor.py
# ==============================================================

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from extractors.utils import nettoyer_valeur, lire_date, extraire_mois_annee


MOTS_DETAIL = ["DETAIL", "DETAILE", "DETAILLE", "DÉTAIL", "DÉTAILLE"]

MOIS_DETECTABLES = [
    "JANVIER", "FEVRIER", "FÉVRIER", "MARS", "AVRIL", "MAI", "JUIN",
    "JUILLET", "AOUT", "AOÛT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE", "DÉCEMBRE",
    "JANV", "FEVR", "FÉVR", "AVRI", "JUIL", "SEPT", "OCTO", "NOVE", "DECE"
]


def _est_feuille_recap(nom: str) -> bool:
    nom_upper = nom.upper().strip()
    for mot in MOTS_DETAIL:
        if mot in nom_upper:
            return False
    return any(mois in nom_upper for mois in MOIS_DETECTABLES)


def _trouver_ligne_entete(lignes: list) -> tuple:
    for i, ligne in enumerate(lignes):
        valeurs = [str(cellule).upper().strip() for cellule in ligne if cellule is not None]
        if "CB" in valeurs and "ESPECE" in valeurs:
            colonnes = {}
            for j, cellule in enumerate(ligne):
                if cellule is None:
                    continue
                cle = str(cellule).upper().strip()
                if cle == "CB" and "cb" not in colonnes:
                    colonnes["cb"] = j
                elif cle == "ESPECE" and "espece" not in colonnes:
                    colonnes["espece"] = j
                elif cle == "PAYPAL" and "paypal" not in colonnes:
                    colonnes["paypal"] = j
                elif cle in ("CHQ", "CHEQUE", "CHÈQUE") and "cheque" not in colonnes:
                    colonnes["cheque"] = j
                elif cle == "VIREMENT" and "virement" not in colonnes:
                    colonnes["virement"] = j
                elif cle == "TOTAL" and "total" not in colonnes:
                    colonnes["total"] = j
            return i, colonnes
    return None, {}


def _lire_feuille_recap(ws, nom_feuille: str) -> list:
    lignes = list(ws.iter_rows(values_only=True))
    idx_entete, colonnes = _trouver_ligne_entete(lignes)
    if idx_entete is None or "total" not in colonnes:
        return []
    enregistrements = []
    for ligne in lignes[idx_entete + 1:]:
        date_val = None
        for k in range(min(4, len(ligne))):
            date_val = lire_date(ligne[k])
            if date_val is not None:
                break
        if date_val is None:
            continue
        def get_col(nom_col):
            if nom_col in colonnes:
                idx = colonnes[nom_col]
                if idx < len(ligne):
                    return nettoyer_valeur(ligne[idx])
            return None
        total = get_col("total")
        if total is None or total == 0:
            continue
        enregistrements.append({
            "date":     date_val,
            "feuille":  nom_feuille,
            "cb":       get_col("cb"),
            "espece":   get_col("espece"),
            "paypal":   get_col("paypal"),
            "cheque":   get_col("cheque"),
            "virement": get_col("virement"),
            "total":    total,
        })
    return enregistrements


def extraire_ca(chemin_fichier: str) -> dict:
    try:
        wb = openpyxl.load_workbook(chemin_fichier, read_only=True, data_only=True)
    except FileNotFoundError:
        return {"donnees": [], "nb_jours": 0, "erreurs": [f"Fichier introuvable : {chemin_fichier}"]}
    except Exception as e:
        return {"donnees": [], "nb_jours": 0, "erreurs": [f"Impossible d'ouvrir : {e}"]}

    toutes_donnees = []
    erreurs = []
    feuilles_lues = 0

    for nom_feuille in wb.sheetnames:
        if not _est_feuille_recap(nom_feuille):
            continue
        try:
            ws = wb[nom_feuille]
            donnees = _lire_feuille_recap(ws, nom_feuille)
            if donnees:
                toutes_donnees.extend(donnees)
                feuilles_lues += 1
        except Exception as e:
            erreurs.append(f"Erreur '{nom_feuille}' : {e}")

    wb.close()
    return {
        "donnees":  toutes_donnees,
        "nb_jours": len(toutes_donnees),
        "erreurs":  erreurs
    }
