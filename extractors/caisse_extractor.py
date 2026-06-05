import sys, os, re, datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from extractors.utils import nettoyer_valeur, lire_date, extraire_mois_annee
from config.mapping import CAISSE_COL_CODE, CAISSE_COL_DEBIT, CAISSE_COL_CREDIT, CAISSE_CODES_CA


def _extraire_date_libelle(libelle, annee_feuille):
    if not libelle or not isinstance(libelle, str):
        return None
    lib = libelle.upper().strip()
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})', lib)
    if m:
        j, mo, an = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if abs(an - annee_feuille) > 1:
            an = annee_feuille
        try:
            return datetime.date(an, mo, j)
        except ValueError:
            pass
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})(?![/\-\.]\d)', lib)
    if m:
        j, mo = int(m.group(1)), int(m.group(2))
        if 1 <= j <= 31 and 1 <= mo <= 12:
            try:
                return datetime.date(annee_feuille, mo, j)
            except ValueError:
                pass
    return None


def _trouver_entete_caisse(lignes):
    mots_code   = [c.upper() for c in CAISSE_COL_CODE]
    mots_debit  = [c.upper() for c in CAISSE_COL_DEBIT]
    mots_credit = [c.upper() for c in CAISSE_COL_CREDIT]
    for i, ligne in enumerate(lignes[:20]):
        vals = [str(c).upper().strip() for c in ligne if c is not None]
        if 'DATE' not in vals:
            continue
        if not any(m in vals for m in mots_debit):
            continue
        colonnes = {}
        for j, cell in enumerate(ligne):
            if cell is None:
                continue
            cle = str(cell).upper().strip()
            if cle in mots_code and 'code' not in colonnes:
                colonnes['code'] = j
            elif cle == 'DATE':
                colonnes['date'] = j
            elif cle == 'FNR' and 'fnr' not in colonnes:
                colonnes['fnr'] = j
            elif cle in ('LIBELLÉS','LIBELLE','LIBELLÉ','LIBELLES') and 'libelle' not in colonnes:
                colonnes['libelle'] = j
            elif cle in mots_debit and 'debit' not in colonnes:
                colonnes['debit'] = j
            elif cle in mots_credit and 'credit' not in colonnes:
                colonnes['credit'] = j
            elif cle == 'SOLDE' and 'solde' not in colonnes:
                colonnes['solde'] = j
        if 'date' in colonnes and 'debit' in colonnes:
            return i, colonnes
    return None, {}


def _est_entree_ca(code, fnr, libelle):
    lib = libelle.upper().strip()
    if re.match(r'CA\s+DU\s+\d{1,2}/\d{1,2}', lib):
        return True
    if re.match(r'CA\s+\d{1,2}/\d{1,2}', lib):
        return True
    if re.match(r"CHIFFRE D['\s]?AFFAIRE[S]?\s+(DU\s+)?\d{1,2}/\d{1,2}", lib):
        return True
    if fnr.upper().strip() == 'CA' and re.search(r'\d{1,2}/\d{1,2}', lib):
        return True
    return False


def _lire_feuille_caisse(ws, nom_feuille, id_caisse):
    lignes = list(ws.iter_rows(values_only=True))
    idx_entete, colonnes = _trouver_entete_caisse(lignes)
    if idx_entete is None:
        return []
    my = extraire_mois_annee(nom_feuille)
    annee_feuille = my[1] if my else datetime.date.today().year
    mouvements = []
    for ligne in lignes[idx_entete + 1:]:
        if not ligne or len(ligne) <= max(colonnes.values(), default=0):
            continue
        def get(nom):
            if nom in colonnes and colonnes[nom] < len(ligne):
                return ligne[colonnes[nom]]
            return None
        code    = str(get('code')    or '').strip()
        fnr     = str(get('fnr')     or '').strip()
        libelle = str(get('libelle') or '').strip()
        debit   = nettoyer_valeur(get('debit'))
        if not _est_entree_ca(code, fnr, libelle):
            continue
        if debit is None or debit <= 0:
            continue
        date_lib = _extraire_date_libelle(libelle, annee_feuille)
        date_col = lire_date(get('date'))
        if date_lib is not None:
            date_val    = date_lib
            source_date = 'libelle'
        elif date_col is not None:
            date_val    = date_col
            source_date = 'colonne'
        else:
            continue
        mouvements.append({
            'caisse':        id_caisse,
            'feuille':       nom_feuille,
            'date':          date_val,
            'date_colonne':  date_col,
            'date_libelle':  date_lib,
            'source_date':   source_date,
            'code':          code,
            'fnr':           fnr,
            'libelle':       libelle,
            'debit':         debit,
            'credit':        nettoyer_valeur(get('credit')),
            'solde':         nettoyer_valeur(get('solde')),
            'est_entree_ca': True,
        })
    return mouvements


def extraire_caisse(chemin_fichier, id_caisse):
    try:
        wb = openpyxl.load_workbook(chemin_fichier, read_only=True, data_only=True)
    except Exception as e:
        return {'mouvements':[],'entrees_ca':[],'nb_mouvements':0,'nb_entrees_ca':0,'erreurs':[str(e)]}
    tous = []
    erreurs = []
    for nom_feuille in wb.sheetnames:
        try:
            ws = wb[nom_feuille]
            mvts = _lire_feuille_caisse(ws, nom_feuille, id_caisse)
            if mvts:
                tous.extend(mvts)
        except Exception as e:
            erreurs.append(f'Erreur {nom_feuille}: {e}')
    wb.close()
    return {
        'mouvements':    tous,
        'entrees_ca':    tous,
        'nb_mouvements': len(tous),
        'nb_entrees_ca': len(tous),
        'erreurs':       erreurs,
    }
