# ==============================================================
#  extractors/cpc_extractor.py
#  Lit le fichier CPC mensuel (Compte de Produits et Charges).
#  Structure attendue : feuilles par mois ("MAI 2025", "JANVIER 2026"...)
#  Chaque feuille : PRODUITS D'EXPLOITATION → lignes → TOTAL
#                   CHARGES D'EXPLOITATION  → lignes → TOTAL
#                   RÉSULTAT D'EXPLOITATION
#                   CNSS / RÉSULTAT NET
# ==============================================================

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from extractors.utils import nettoyer_valeur, extraire_mois_annee
from config.mapping import (
    CPC_MOT_PRODUITS, CPC_MOT_CHARGES, CPC_MOT_RESULTAT,
    CPC_MOT_CNSS, CPC_MOT_RESULTAT_NET
)


def _lire_feuille_cpc(ws, nom_feuille: str) -> dict:
    """
    Parse une feuille CPC et retourne un dictionnaire structuré.
    
    Retourne :
    {
        "mois": int, "annee": int, "feuille": str,
        "produits": {"LOCATION TERRAIN": 190652, "COURS": 66150, ...},
        "charges":  {"SALAIRES W": 36000, "LOYER": 15000, ...},
        "total_produits": float,
        "total_charges":  float,
        "resultat_exploitation": float,
        "cnss": float,
        "resultat_net": float,
        "ebitda": float,       # = résultat d'exploitation (approché)
        "marge_ebitda": float, # en %
    }
    """
    mois_annee = extraire_mois_annee(nom_feuille)
    if mois_annee is None:
        return None
    mois, annee = mois_annee
    
    # Lire toutes les lignes non vides
    lignes = [
        row for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]
    
    produits       = {}
    charges        = {}
    total_produits = None
    total_charges  = None
    resultat_expl  = None
    cnss           = None
    resultat_net   = None
    section        = None  # "PRODUITS" ou "CHARGES"
    
    for ligne in lignes:
        # Extraire les cellules non nulles
        cellules = [(i, c) for i, c in enumerate(ligne) if c is not None]
        if len(cellules) < 2:
            continue
        
        # Chercher le premier texte de la ligne
        label = None
        montant = None
        label_trouve = False
        
        for i, cellule in cellules:
            if isinstance(cellule, str) and not label_trouve:
                label = cellule.upper().strip()
                label_trouve = True
            elif label_trouve and montant is None:
                # Le montant est le premier nombre après le label
                if isinstance(cellule, (int, float)):
                    montant = nettoyer_valeur(cellule)
                elif isinstance(cellule, str):
                    m = nettoyer_valeur(cellule)
                    if m is not None:
                        montant = m
        
        if label is None:
            continue
        
        # ── Détecter les changements de section ──────────────────
        if CPC_MOT_PRODUITS.upper() in label:
            section = "PRODUITS"
            continue
        
        if CPC_MOT_CHARGES.upper() in label:
            section = "CHARGES"
            continue
        
        # ── Détecter les lignes spéciales ────────────────────────
        if label.startswith("TOTAL") and montant is not None:
            if section == "PRODUITS" and total_produits is None:
                total_produits = montant
            elif section == "CHARGES" and total_charges is None:
                total_charges = montant
            continue
        
        if CPC_MOT_RESULTAT.upper() in label and montant is not None:
            resultat_expl = montant
            continue
        
        if label == CPC_MOT_CNSS.upper() and montant is not None:
            cnss = montant
            continue
        
        if label == CPC_MOT_RESULTAT_NET.upper() and montant is not None and resultat_expl is not None:
            # C'est le résultat net (après CNSS et TVA)
            resultat_net = montant
            continue
        
        # ── Enregistrer dans la bonne section ────────────────────
        if montant is not None and montant > 0 and section in ("PRODUITS", "CHARGES"):
            # Éviter les doublons en ajoutant
            if section == "PRODUITS":
                produits[label] = produits.get(label, 0) + montant
            elif section == "CHARGES":
                charges[label] = charges.get(label, 0) + montant
    
    # Recalcul des totaux si manquants dans le fichier
    if total_produits is None and produits:
        total_produits = sum(v for v in produits.values())
    if total_charges is None and charges:
        total_charges = sum(v for v in charges.values())
    
    # EBITDA approché = Résultat d'exploitation
    ebitda = resultat_expl
    marge_ebitda = None
    if ebitda is not None and total_produits and total_produits > 0:
        marge_ebitda = round(ebitda / total_produits * 100, 2)
    
    return {
        "mois":                   mois,
        "annee":                  annee,
        "feuille":                nom_feuille,
        "produits":               produits,
        "charges":                charges,
        "total_produits":         total_produits,
        "total_charges":          total_charges,
        "resultat_exploitation":  resultat_expl,
        "cnss":                   cnss,
        "resultat_net":           resultat_net,
        "ebitda":                 ebitda,
        "marge_ebitda":           marge_ebitda,
    }


def extraire_cpc(chemin_fichier: str) -> dict:
    """
    FONCTION PRINCIPALE — Extrait tous les CPC mensuels.
    
    Paramètre :
        chemin_fichier : chemin complet vers le fichier Excel CPC
    
    Retourne :
        {
            "mensuel":    liste de dicts CPC (triés par date),
            "nb_mois":    int,
            "erreurs":    liste de messages
        }
    """
    print(f"\n📂 Lecture du CPC...")
    print(f"   Fichier : {os.path.basename(chemin_fichier)}")
    
    try:
        wb = openpyxl.load_workbook(chemin_fichier, read_only=True, data_only=True)
    except FileNotFoundError:
        return {"mensuel": [], "nb_mois": 0, "erreurs": [f"Fichier introuvable : {chemin_fichier}"]}
    except Exception as e:
        return {"mensuel": [], "nb_mois": 0, "erreurs": [f"Impossible d'ouvrir le fichier : {e}"]}
    
    tous_mois = []
    erreurs = []
    
    for nom_feuille in wb.sheetnames:
        try:
            ws = wb[nom_feuille]
            result = _lire_feuille_cpc(ws, nom_feuille)
            if result is not None:
                ca = result.get("total_produits")
                ch = result.get("total_charges")
                ca_str = f"{ca:,.0f}" if ca else "?"
                ch_str = f"{ch:,.0f}" if ch else "?"
                print(f"   ✅ {nom_feuille.strip()} → CA={ca_str} DH / Charges={ch_str} DH")
                tous_mois.append(result)
        except Exception as e:
            msg = f"Erreur feuille '{nom_feuille}' : {e}"
            erreurs.append(msg)
            print(f"   ❌ {msg}")
    
    wb.close()
    
    # Trier par année puis mois
    tous_mois.sort(key=lambda x: (x["annee"], x["mois"]))
    
    print(f"   → Total : {len(tous_mois)} mois extraits")
    
    return {
        "mensuel":  tous_mois,
        "nb_mois":  len(tous_mois),
        "erreurs":  erreurs
    }
