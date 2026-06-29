# ==============================================================
#  extractors/planning_extractor.py
#  Calcule le taux de remplissage reel depuis le planning de
#  reservation des terrains.
#
#  Regle de lecture (simple et fiable) :
#   - RESERVE : la cellule contient du texte (nom, prix, note)
#   - LIBRE   : la cellule est vide
#
#  Taux = creneaux reserves / total creneaux
# ==============================================================

import os
import datetime
import openpyxl


def _est_reserve(cell) -> bool:
    """RESERVE uniquement si la cellule contient du texte."""
    return cell.value is not None and str(cell.value).strip() != ""


def extraire_taux_remplissage(chemin_fichier: str) -> dict:
    """Analyse le planning et retourne le taux de remplissage par jour."""
    try:
        wb = openpyxl.load_workbook(chemin_fichier, data_only=True)
    except FileNotFoundError:
        return {"jours": [], "taux_moyen": None, "total_reserves": 0,
                "total_creneaux": 0, "erreurs": [f"Fichier introuvable : {chemin_fichier}"]}
    except Exception as e:
        return {"jours": [], "taux_moyen": None, "total_reserves": 0,
                "total_creneaux": 0, "erreurs": [f"Impossible d'ouvrir : {e}"]}

    jours = []
    erreurs = []

    for ws in wb.worksheets:
        titre = ws.title.lower()
        if any(mot in titre for mot in ["coach", "suivi", "heure", "fixe"]):
            continue
        try:
            jours.extend(_analyser_feuille(ws))
        except Exception as e:
            erreurs.append(f"Erreur feuille '{ws.title}' : {e}")

    wb.close()

    jours_dict = {j["date"]: j for j in jours}
    jours_final = sorted(jours_dict.values(), key=lambda x: x["date"])

    total_res = sum(j["reserves"] for j in jours_final)
    total_cre = sum(j["total"]    for j in jours_final)
    taux_moyen = (total_res / total_cre * 100) if total_cre > 0 else None

    return {
        "jours":          jours_final,
        "taux_moyen":     taux_moyen,
        "total_reserves": total_res,
        "total_creneaux": total_cre,
        "erreurs":        erreurs,
    }


def _analyser_feuille(ws) -> list:
    rows_all = list(ws.iter_rows(values_only=True))

    ligne_dates = None
    for i, row in enumerate(rows_all[:10]):
        if sum(1 for c in row if isinstance(c, datetime.datetime)) >= 3:
            ligne_dates = i
            break
    if ligne_dates is None:
        return []

    dates_cols = [(j, c.date()) for j, c in enumerate(rows_all[ligne_dates])
                  if isinstance(c, datetime.datetime)]
    if not dates_cols:
        return []

    col_heure = dates_cols[0][0] - 1

    lignes_horaires = []
    for r in range(ligne_dates, len(rows_all)):
        if col_heure < len(rows_all[r]):
            if isinstance(rows_all[r][col_heure], datetime.time):
                lignes_horaires.append(r)
    if not lignes_horaires:
        return []

    resultats = []
    for idx, (col_date, date) in enumerate(dates_cols):
        if idx + 1 < len(dates_cols):
            col_fin = dates_cols[idx + 1][0] - 1
        else:
            col_fin = col_date + 11

        nb_reserves = 0
        nb_total = 0
        for row_idx in lignes_horaires:
            for col in range(col_date, col_fin):
                cell = ws.cell(row=row_idx + 1, column=col + 1)
                nb_total += 1
                if _est_reserve(cell):
                    nb_reserves += 1

        if nb_total > 0:
            resultats.append({
                "date":     date,
                "reserves": nb_reserves,
                "libres":   nb_total - nb_reserves,
                "total":    nb_total,
                "taux":     nb_reserves / nb_total * 100,
            })

    return resultats
